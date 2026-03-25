"""
Lab 02: Multiple Linear Regression, Cross Validation & Polynomial Regression
==============================================================================
Implements from scratch:
    Part A - Linear regression with multiple variables (CCPP dataset)
    Part B - 5-fold cross validation
    Part C - Polynomial regression with a single variable (data_02b.csv)

All algorithms use vectorized NumPy operations and gradient descent.
"""

import numpy as np
import matplotlib.pyplot as plt
import os

# ─────────────────────────── Configuration ──────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLOTS_DIR = os.path.join(BASE_DIR, "plots")
os.makedirs(PLOTS_DIR, exist_ok=True)

CCPP_CONFIG = {
    "file_path": os.path.join(BASE_DIR, "CCPP", "CCPP", "Folds5x2_pp.xlsx"),
    "feature_names": [
        "AT (Temperature)",
        "V (Exhaust Vacuum)",
        "AP (Ambient Pressure)",
        "RH (Relative Humidity)",
    ],
    "target_name": "PE (Net Energy Output)",
    "learning_rate": 0.01,
    "iterations": 1500,
    "train_ratio": 0.8,
}

POLY_CONFIG = {
    "file_path": os.path.join(BASE_DIR, "data_02b.csv"),
    "learning_rate": 0.01,
    "iterations": 2000,
    "degrees": [1, 2, 3],
    "train_ratio": 0.8,
}


# ═══════════════════════════════════════════════════════════════════════════
#  CORE ML FUNCTIONS (reusable across all parts)
# ═══════════════════════════════════════════════════════════════════════════


def load_ccpp_data(filepath):
    """
    Load the CCPP dataset from an Excel file (first sheet).

    Returns:
        X : (m, 4) array of features [AT, V, AP, RH]
        y : (m,)   array of target values [PE]
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        data.append([float(v) for v in row])
    wb.close()

    data = np.array(data)
    X = data[:, :4]
    y = data[:, 4]
    print(f"[INFO] Loaded CCPP data: {X.shape[0]} samples, {X.shape[1]} features")
    return X, y


def load_csv_data(filepath):
    """
    Load a CSV file with no header. Returns feature(s) and target.

    Returns:
        x : (m,) feature array
        y : (m,) target array
    """
    data = np.loadtxt(filepath, delimiter=",")
    x = data[:, 0]
    y = data[:, 1]
    print(f"[INFO] Loaded {x.shape[0]} samples from {os.path.basename(filepath)}")
    return x, y


def train_test_split(X, y, train_ratio=0.8, seed=42):
    """
    Shuffle and split data into training and validation sets.

    Returns:
        X_train, y_train, X_val, y_val
    """
    m = X.shape[0]
    np.random.seed(seed)
    indices = np.random.permutation(m)
    split = int(m * train_ratio)

    train_idx = indices[:split]
    val_idx = indices[split:]

    return X[train_idx], y[train_idx], X[val_idx], y[val_idx]


def normalize_features(X_train, X_val=None):
    """
    Z-score normalization: (x - mean) / std.

    Computes statistics from X_train only, applies to both sets.

    Returns:
        X_train_norm, X_val_norm (or None), norm_params dict
    """
    mu = np.mean(X_train, axis=0)
    sigma = np.std(X_train, axis=0)
    sigma[sigma == 0] = 1  # avoid division by zero

    X_train_norm = (X_train - mu) / sigma
    X_val_norm = (X_val - mu) / sigma if X_val is not None else None

    norm_params = {"mean": mu, "std": sigma}
    return X_train_norm, X_val_norm, norm_params


def add_bias(X):
    """Prepend a column of ones (bias term) to the design matrix."""
    m = X.shape[0]
    return np.column_stack((np.ones(m), X))


def compute_cost(X, y, theta):
    """
    Mean Squared Error cost: J(θ) = (1 / 2m) * ||Xθ - y||².

    Parameters:
        X     : (m, n) design matrix (with bias column)
        y     : (m, 1) target vector
        theta : (n, 1) parameter vector

    Returns:
        cost : scalar
    """
    m = X.shape[0]
    errors = X @ theta - y
    return (1.0 / (2 * m)) * np.sum(errors**2)


def gradient_descent(X, y, theta, alpha, num_iters, X_val=None, y_val=None):
    """
    Batch gradient descent with optional validation tracking.

    Update rule: θ := θ - (α / m) * Xᵀ(Xθ - y)

    Returns:
        theta        : (n, 1) learned parameters
        train_costs  : list of training cost per iteration
        val_costs    : list of validation cost per iteration (empty if no val set)
    """
    m = X.shape[0]
    train_costs = []
    val_costs = []

    for _ in range(num_iters):
        predictions = X @ theta
        errors = predictions - y
        gradient = (1.0 / m) * (X.T @ errors)
        theta = theta - alpha * gradient

        train_costs.append(compute_cost(X, y, theta))
        if X_val is not None:
            val_costs.append(compute_cost(X_val, y_val, theta))

    return theta, train_costs, val_costs


def train_model(X_train, y_train, X_val=None, y_val=None, alpha=0.01, num_iters=1000):
    """
    Train a linear regression model using gradient descent.

    Parameters:
        X_train   : (m, n) design matrix (with bias)
        y_train   : (m, 1) target
        X_val     : (m_val, n) validation design matrix (optional)
        y_val     : (m_val, 1) validation target (optional)
        alpha     : learning rate
        num_iters : number of iterations

    Returns:
        theta, train_costs, val_costs
    """
    n = X_train.shape[1]
    theta = np.zeros((n, 1))

    theta, train_costs, val_costs = gradient_descent(
        X_train, y_train, theta, alpha, num_iters, X_val, y_val
    )

    print(f"[TRAIN] Final training cost   = {train_costs[-1]:.6f}")
    if val_costs:
        print(f"[TRAIN] Final validation cost = {val_costs[-1]:.6f}")

    return theta, train_costs, val_costs


def k_fold_split(X, y, k=5, seed=42):
    """
    Generate k-fold cross-validation splits.

    Yields:
        (X_train, y_train, X_val, y_val) for each fold
    """
    m = X.shape[0]
    np.random.seed(seed)
    indices = np.random.permutation(m)
    fold_size = m // k

    for i in range(k):
        val_start = i * fold_size
        val_end = val_start + fold_size if i < k - 1 else m
        val_idx = indices[val_start:val_end]
        train_idx = np.concatenate([indices[:val_start], indices[val_end:]])
        yield X[train_idx], y[train_idx], X[val_idx], y[val_idx]


def build_polynomial_features(x, degree):
    """
    Create polynomial feature matrix [x, x², ..., x^d].

    Parameters:
        x      : (m,) raw feature vector
        degree : polynomial degree d

    Returns:
        X_poly : (m, d) feature matrix (no bias column)
    """
    m = x.shape[0]
    X_poly = np.zeros((m, degree))
    for d in range(1, degree + 1):
        X_poly[:, d - 1] = x**d
    return X_poly


# ═══════════════════════════════════════════════════════════════════════════
#  PLOTTING UTILITIES
# ═══════════════════════════════════════════════════════════════════════════


def plot_feature_vs_target(X, y, feature_names, target_name, save_prefix=None):
    """Plot each feature against the target variable (grid of subplots)."""
    n_features = X.shape[1]
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    axes = axes.flatten()

    for i in range(n_features):
        ax = axes[i]
        ax.scatter(X[:, i], y, s=5, alpha=0.3, edgecolors="none")
        ax.set_xlabel(feature_names[i])
        ax.set_ylabel(target_name)
        ax.set_title(f"{feature_names[i]} vs {target_name}")
        ax.grid(True, alpha=0.3)

    plt.suptitle("Individual Features vs Target Variable", fontsize=14, y=1.01)
    plt.tight_layout()
    if save_prefix:
        path = os.path.join(PLOTS_DIR, f"{save_prefix}_features_vs_target.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"[INFO] Plot saved: {path}")
    plt.show()


def plot_error_curves(train_costs, val_costs, title="Error Curves", save_as=None):
    """Plot training and validation error vs. iterations."""
    plt.figure(figsize=(10, 6))
    iters = range(1, len(train_costs) + 1)
    plt.plot(iters, train_costs, label="Training Error", linewidth=1.5)
    if val_costs:
        plt.plot(iters, val_costs, label="Validation Error", linewidth=1.5)
    plt.xlabel("Iteration")
    plt.ylabel("Cost J(θ)")
    plt.title(title)
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    if save_as:
        path = os.path.join(PLOTS_DIR, save_as)
        plt.savefig(path, dpi=150)
        print(f"[INFO] Plot saved: {path}")
    plt.show()


def plot_scatter(x, y, xlabel="x", ylabel="y", title="Data", save_as=None):
    """Simple scatter plot."""
    plt.figure(figsize=(8, 5))
    plt.scatter(x, y, s=15, alpha=0.6)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    if save_as:
        path = os.path.join(PLOTS_DIR, save_as)
        plt.savefig(path, dpi=150)
        print(f"[INFO] Plot saved: {path}")
    plt.show()


def print_parameters(theta, label=""):
    """Pretty-print learned parameters."""
    print(f"\n{'=' * 55}")
    print(f"  LEARNED PARAMETERS {label}")
    print(f"{'=' * 55}")
    for i, val in enumerate(theta.flatten()):
        print(f"  theta_{i:<2d} = {val:>14.6f}")
    print(f"{'=' * 55}\n")


def _trow(*cols, widths):
    """Format a table row by left-justifying each column to its width."""
    return "  " + "".join(str(c).ljust(w) for c, w in zip(cols, widths))


def _tsep(widths, char="─"):
    """Print a horizontal separator matching total column width."""
    return "  " + char * sum(widths)


# ═══════════════════════════════════════════════════════════════════════════
#  PART A: LINEAR REGRESSION WITH MULTIPLE VARIABLES
# ═══════════════════════════════════════════════════════════════════════════


def run_part_a():
    """Part A: Multi-variable linear regression on the CCPP dataset."""
    cfg = CCPP_CONFIG
    print("\n" + "=" * 65)
    print("  PART A: LINEAR REGRESSION WITH MULTIPLE VARIABLES (CCPP)")
    print("=" * 65 + "\n")

    # ── 1. Load data ──
    X_raw, y_raw = load_ccpp_data(cfg["file_path"])

    # ── 2. Plot individual features vs target ──
    plot_feature_vs_target(
        X_raw, y_raw, cfg["feature_names"], cfg["target_name"], save_prefix="partA"
    )

    # ── 3. Train/validation split ──
    X_train_raw, y_train_raw, X_val_raw, y_val_raw = train_test_split(
        X_raw, y_raw, train_ratio=cfg["train_ratio"]
    )
    y_train = y_train_raw.reshape(-1, 1)
    y_val = y_val_raw.reshape(-1, 1)

    # ━━━━━━━━━━  WITHOUT normalization  ━━━━━━━━━━
    print("\n--- Without Feature Normalization ---")
    X_train_nb = add_bias(X_train_raw)
    X_val_nb = add_bias(X_val_raw)

    # Use a smaller learning rate without normalization to avoid divergence
    alpha_no_norm = 1e-7
    theta_no_norm, train_costs_nn, val_costs_nn = train_model(
        X_train_nb,
        y_train,
        X_val_nb,
        y_val,
        alpha=alpha_no_norm,
        num_iters=cfg["iterations"],
    )
    best_val_nn = min(val_costs_nn)
    print(f"  Best validation cost (no norm)  = {best_val_nn:.6f}")
    print_parameters(theta_no_norm, "(without normalization)")

    plot_error_curves(
        train_costs_nn,
        val_costs_nn,
        title="Part A – Error Curves (Without Normalization)",
        save_as="partA_error_no_norm.png",
    )

    # ━━━━━━━━━━  WITH normalization  ━━━━━━━━━━
    print("\n--- With Feature Normalization ---")
    X_train_norm, X_val_norm, norm_params = normalize_features(X_train_raw, X_val_raw)
    X_train_n = add_bias(X_train_norm)
    X_val_n = add_bias(X_val_norm)

    theta_norm, train_costs_n, val_costs_n = train_model(
        X_train_n,
        y_train,
        X_val_n,
        y_val,
        alpha=cfg["learning_rate"],
        num_iters=cfg["iterations"],
    )
    best_val_n = min(val_costs_n)
    best_train_n = min(train_costs_n)
    print(f"  Best training cost   (norm)  = {best_train_n:.6f}")
    print(f"  Best validation cost (norm)  = {best_val_n:.6f}")
    print_parameters(theta_norm, "(with normalization)")

    plot_error_curves(
        train_costs_n,
        val_costs_n,
        title="Part A – Error Curves (With Normalization)",
        save_as="partA_error_norm.png",
    )

    # ── Comparison summary ──
    print("\n" + "─" * 55)
    print("  NORMALIZATION COMPARISON")
    print("─" * 55)
    print(f"  Without normalization (lr={alpha_no_norm}):")
    print(f"    Final train cost = {train_costs_nn[-1]:.6f}")
    print(f"    Final val   cost = {val_costs_nn[-1]:.6f}")
    print(f"  With normalization    (lr={cfg['learning_rate']}):")
    print(f"    Final train cost = {train_costs_n[-1]:.6f}")
    print(f"    Final val   cost = {val_costs_n[-1]:.6f}")
    print("─" * 55 + "\n")

    return theta_norm, norm_params


# ═══════════════════════════════════════════════════════════════════════════
#  PART B: 5-FOLD CROSS VALIDATION
# ═══════════════════════════════════════════════════════════════════════════


def run_part_b():
    """Part B: 5-fold cross validation on the CCPP dataset."""
    cfg = CCPP_CONFIG
    print("\n" + "=" * 65)
    print("  PART B: 5-FOLD CROSS VALIDATION (CCPP)")
    print("=" * 65 + "\n")

    X_raw, y_raw = load_ccpp_data(cfg["file_path"])

    fold_results = []
    all_thetas = []

    for fold_idx, (X_tr, y_tr, X_va, y_va) in enumerate(
        k_fold_split(X_raw, y_raw, k=5)
    ):
        print(f"\n--- Fold {fold_idx + 1} / 5 ---")

        # Normalize using training fold statistics
        X_tr_norm, X_va_norm, _ = normalize_features(X_tr, X_va)
        X_tr_b = add_bias(X_tr_norm)
        X_va_b = add_bias(X_va_norm)

        y_tr_v = y_tr.reshape(-1, 1)
        y_va_v = y_va.reshape(-1, 1)

        theta, train_costs, val_costs = train_model(
            X_tr_b,
            y_tr_v,
            X_va_b,
            y_va_v,
            alpha=cfg["learning_rate"],
            num_iters=cfg["iterations"],
        )

        fold_results.append(
            {
                "fold": fold_idx + 1,
                "train_cost": train_costs[-1],
                "val_cost": val_costs[-1],
            }
        )
        all_thetas.append(theta)

    # ── Summary table ──
    print("\n" + "=" * 55)
    print("  5-FOLD CROSS VALIDATION RESULTS")
    print("=" * 55)
    print(f"  {'Fold':<6} {'Train Cost':<16} {'Val Cost':<16}")
    print("  " + "─" * 40)
    for r in fold_results:
        print(f"  {r['fold']:<6} {r['train_cost']:<16.6f} {r['val_cost']:<16.6f}")

    train_costs_all = [r["train_cost"] for r in fold_results]
    val_costs_all = [r["val_cost"] for r in fold_results]
    print("  " + "─" * 40)
    print(
        f"  {'Mean':<6} {np.mean(train_costs_all):<16.6f} {np.mean(val_costs_all):<16.6f}"
    )
    print(
        f"  {'Std':<6} {np.std(train_costs_all):<16.6f} {np.std(val_costs_all):<16.6f}"
    )
    print("=" * 55)

    # Best fold
    best_fold_idx = np.argmin(val_costs_all)
    best_theta = all_thetas[best_fold_idx]
    print(
        f"\n  Best fold: {best_fold_idx + 1}  "
        f"(val cost = {val_costs_all[best_fold_idx]:.6f})"
    )
    print_parameters(best_theta, f"(Best – Fold {best_fold_idx + 1})")

    # Bar plot of fold validation errors
    plt.figure(figsize=(8, 5))
    folds_labels = [f"Fold {i+1}" for i in range(5)]
    bars = plt.bar(folds_labels, val_costs_all, color="steelblue", edgecolor="black")
    bars[best_fold_idx].set_color("green")
    plt.axhline(
        y=np.mean(val_costs_all),
        color="red",
        linestyle="--",
        label=f"Mean = {np.mean(val_costs_all):.4f}",
    )
    plt.xlabel("Fold")
    plt.ylabel("Validation Cost")
    plt.title("Part B – 5-Fold Cross Validation Errors")
    plt.legend()
    plt.grid(True, alpha=0.3, axis="y")
    plt.tight_layout()
    path = os.path.join(PLOTS_DIR, "partB_cv_errors.png")
    plt.savefig(path, dpi=150)
    print(f"[INFO] Plot saved: {path}")
    plt.show()

    return best_theta


# ═══════════════════════════════════════════════════════════════════════════
#  PART C: POLYNOMIAL REGRESSION
# ═══════════════════════════════════════════════════════════════════════════


def run_part_c():
    """Part C: Polynomial regression on data_02b.csv."""
    cfg = POLY_CONFIG
    print("\n" + "=" * 65)
    print("  PART C: POLYNOMIAL REGRESSION (data_02b.csv)")
    print("=" * 65 + "\n")

    # ── 1. Load and plot data ──
    x_raw, y_raw = load_csv_data(cfg["file_path"])

    plot_scatter(
        x_raw,
        y_raw,
        xlabel="Feature",
        ylabel="Target",
        title="Part C – Raw Data (data_02b.csv)",
        save_as="partC_raw_data.png",
    )

    # ── 2. Train/validation split ──
    # Reshape for compatibility with split function
    X_1d = x_raw.reshape(-1, 1)
    X_train_raw, y_train_raw, X_val_raw, y_val_raw = train_test_split(
        X_1d, y_raw, train_ratio=cfg["train_ratio"]
    )
    x_train = X_train_raw.flatten()
    x_val = X_val_raw.flatten()
    y_train = y_train_raw.reshape(-1, 1)
    y_val = y_val_raw.reshape(-1, 1)

    # ── 3. Polynomial regression for d = 1, 2, 3 ──
    results = {}  # degree -> {theta, train_costs, val_costs, norm_params, val_error}

    for d in cfg["degrees"]:
        print(f"\n{'─' * 45}")
        print(f"  Polynomial Degree d = {d}")
        print(f"{'─' * 45}")

        # Build polynomial features
        X_train_poly = build_polynomial_features(x_train, d)
        X_val_poly = build_polynomial_features(x_val, d)

        # Normalize
        X_train_norm, X_val_norm, norm_params = normalize_features(
            X_train_poly, X_val_poly
        )

        # Add bias
        X_train_b = add_bias(X_train_norm)
        X_val_b = add_bias(X_val_norm)

        # Train
        theta, train_costs, val_costs = train_model(
            X_train_b,
            y_train,
            X_val_b,
            y_val,
            alpha=cfg["learning_rate"],
            num_iters=cfg["iterations"],
        )

        results[d] = {
            "theta": theta,
            "train_costs": train_costs,
            "val_costs": val_costs,
            "norm_params": norm_params,
            "val_error": val_costs[-1],
            "train_error": train_costs[-1],
        }

        print_parameters(theta, f"(degree {d})")

    # ── 4. Find best degree ──
    best_d = min(results, key=lambda d: results[d]["val_error"])
    print(f"\n{'=' * 55}")
    print(
        f"  BEST DEGREE: d = {best_d}  "
        f"(val cost = {results[best_d]['val_error']:.6f})"
    )
    print(f"{'=' * 55}")
    print_parameters(results[best_d]["theta"], f"(BEST, degree {best_d})")

    # ── 5. Plot all three fitted curves on one graph ──
    plt.figure(figsize=(10, 7))
    plt.scatter(x_raw, y_raw, s=15, alpha=0.4, color="gray", label="Data")

    x_line = np.linspace(x_raw.min(), x_raw.max(), 300)
    colors = {1: "blue", 2: "green", 3: "red"}

    for d in cfg["degrees"]:
        r = results[d]
        X_line_poly = build_polynomial_features(x_line, d)
        # Normalize using training statistics
        X_line_norm = (X_line_poly - r["norm_params"]["mean"]) / r["norm_params"]["std"]
        X_line_b = add_bias(X_line_norm)
        y_pred = X_line_b @ r["theta"]

        plt.plot(
            x_line,
            y_pred.flatten(),
            color=colors[d],
            linewidth=2,
            label=f"d={d} (val={r['val_error']:.4f})",
        )

    plt.xlabel("Feature")
    plt.ylabel("Target")
    plt.title("Part C – Polynomial Regression Fits (d = 1, 2, 3)")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    path = os.path.join(PLOTS_DIR, "partC_poly_fits.png")
    plt.savefig(path, dpi=150)
    print(f"[INFO] Plot saved: {path}")
    plt.show()

    # ── 6. Bar plot of validation errors ──
    plt.figure(figsize=(8, 5))
    degrees_str = [f"d = {d}" for d in cfg["degrees"]]
    val_errors = [results[d]["val_error"] for d in cfg["degrees"]]
    bar_colors = ["green" if d == best_d else "steelblue" for d in cfg["degrees"]]
    plt.bar(degrees_str, val_errors, color=bar_colors, edgecolor="black")
    plt.xlabel("Polynomial Degree")
    plt.ylabel("Validation Cost")
    plt.title("Part C – Validation Errors for Different Polynomial Degrees")
    plt.grid(True, alpha=0.3, axis="y")
    plt.tight_layout()
    path = os.path.join(PLOTS_DIR, "partC_val_errors_bar.png")
    plt.savefig(path, dpi=150)
    print(f"[INFO] Plot saved: {path}")
    plt.show()

    # ── 7. Training & validation error curves for all degrees ──
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    for idx, d in enumerate(cfg["degrees"]):
        ax = axes[idx]
        r = results[d]
        iters = range(1, len(r["train_costs"]) + 1)
        ax.plot(iters, r["train_costs"], label="Train", linewidth=1.2)
        ax.plot(iters, r["val_costs"], label="Validation", linewidth=1.2)
        ax.set_xlabel("Iteration")
        ax.set_ylabel("Cost J(θ)")
        ax.set_title(f"d = {d}")
        ax.legend()
        ax.grid(True, alpha=0.3)

    plt.suptitle("Part C – Error Curves for Polynomial Degrees", fontsize=14)
    plt.tight_layout()
    path = os.path.join(PLOTS_DIR, "partC_error_curves_all.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    print(f"[INFO] Plot saved: {path}")
    plt.show()

    return results


# ═══════════════════════════════════════════════════════════════════════════
#  SECTION 9: DETAILED RESULTS — VALUES MAPPED TO DATASET & COMPARISON TABLES
# ═══════════════════════════════════════════════════════════════════════════


def print_result_tables(theta_a, theta_b, poly_results):
    """
    Print 6 formatted comparison tables:
      Table 1 — Part A: learned θ mapped to CCPP feature names
      Table 2 — Part A: with vs without normalization cost comparison
      Table 3 — Part B: per-fold CV results + best-fold θ mapped to features
      Table 4 — Part A vs Part B: parameter-by-parameter comparison
      Table 5 — Part C: polynomial degree comparison + θ per degree
      Table 6 — Overall summary of all experiments
    """
    cfg = CCPP_CONFIG
    feature_names = cfg["feature_names"]
    target_name = cfg["target_name"]
    param_labels = ["Intercept (bias)"] + feature_names

    # ── TABLE 1: Part A — theta mapped to features ──────────────────────────
    print("\n" + "=" * 70)
    print("  TABLE 1 — PART A: Learned Parameters (CCPP, With Normalization)")
    print("=" * 70)
    w = [6, 30, 18, 18]
    print(
        _trow(
            "θ", "Feature / Variable", "Value (norm.space)", "Interpretation", widths=w
        )
    )
    print(_tsep(w))
    for i, (lbl, val) in enumerate(zip(param_labels, theta_a.flatten())):
        interp = (
            "baseline PE when all features = their mean"
            if i == 0
            else (
                "↑ temp → ↓ output"
                if i == 1
                else (
                    "↑ vacuum → ↓ output"
                    if i == 2
                    else (
                        "↑ pressure → slight ↑ output"
                        if i == 3
                        else "↑ humidity → ↓ output"
                    )
                )
            )
        )
        print(_trow(f"θ{i}", lbl, f"{val:+.6f}", interp, widths=w))
    print(_tsep(w))
    print(f"  Target variable: {target_name}")

    # ── TABLE 2: Part A — with vs without normalization ──────────────────────
    print("\n\n" + "=" * 70)
    print("  TABLE 2 — PART A: With vs Without Normalization")
    print("=" * 70)
    w2 = [30, 12, 12, 14, 16]
    print(
        _trow("Setting", "LR (α)", "Iters", "Final Train J", "Final Val J", widths=w2)
    )
    print(_tsep(w2))

    X_r, y_r = load_ccpp_data(cfg["file_path"])
    X_tr_r, y_tr_r, X_v_r, y_v_r = train_test_split(
        X_r, y_r, train_ratio=cfg["train_ratio"]
    )
    y_tr_r = y_tr_r.reshape(-1, 1)
    y_v_r = y_v_r.reshape(-1, 1)

    _, tc_nn, vc_nn = train_model(
        add_bias(X_tr_r),
        y_tr_r,
        add_bias(X_v_r),
        y_v_r,
        alpha=1e-7,
        num_iters=cfg["iterations"],
    )
    Xn_tr, Xn_v, _ = normalize_features(X_tr_r, X_v_r)
    _, tc_n, vc_n = train_model(
        add_bias(Xn_tr),
        y_tr_r,
        add_bias(Xn_v),
        y_v_r,
        alpha=cfg["learning_rate"],
        num_iters=cfg["iterations"],
    )

    print(
        _trow(
            "Without normalization",
            "1e-7",
            cfg["iterations"],
            f"{tc_nn[-1]:.4f}",
            f"{vc_nn[-1]:.4f}",
            widths=w2,
        )
    )
    print(
        _trow(
            "With Z-score normalization",
            "0.01",
            cfg["iterations"],
            f"{tc_n[-1]:.4f}",
            f"{vc_n[-1]:.4f}",
            widths=w2,
        )
    )
    print(_tsep(w2))
    improvement = ((vc_nn[-1] - vc_n[-1]) / vc_nn[-1]) * 100
    print(f"  → Normalization reduces validation cost by {improvement:.1f}%")

    # ── TABLE 3: Part B — per-fold CV results ────────────────────────────────
    print("\n\n" + "=" * 70)
    print("  TABLE 3 — PART B: 5-Fold Cross Validation Results (CCPP)")
    print("=" * 70)

    fold_train_costs, fold_val_costs, fold_thetas = [], [], []
    X_cv, y_cv = load_ccpp_data(cfg["file_path"])
    for X_tr_f, y_tr_f, X_va_f, y_va_f in k_fold_split(X_cv, y_cv, k=5):
        Xn_tr_f, Xn_va_f, _ = normalize_features(X_tr_f, X_va_f)
        th_f, tc_f, vc_f = train_model(
            add_bias(Xn_tr_f),
            y_tr_f.reshape(-1, 1),
            add_bias(Xn_va_f),
            y_va_f.reshape(-1, 1),
            alpha=cfg["learning_rate"],
            num_iters=cfg["iterations"],
        )
        fold_train_costs.append(tc_f[-1])
        fold_val_costs.append(vc_f[-1])
        fold_thetas.append(th_f)

    best_fold = int(np.argmin(fold_val_costs))
    w3 = [8, 16, 16, 10]
    print(_trow("Fold", "Train Cost J", "Val Cost J", "Status", widths=w3))
    print(_tsep(w3))
    for i in range(5):
        status = "★ BEST" if i == best_fold else ""
        print(
            _trow(
                f"  {i+1}",
                f"{fold_train_costs[i]:.6f}",
                f"{fold_val_costs[i]:.6f}",
                status,
                widths=w3,
            )
        )
    print(_tsep(w3))
    print(
        _trow(
            "Mean",
            f"{np.mean(fold_train_costs):.6f}",
            f"{np.mean(fold_val_costs):.6f}",
            "",
            widths=w3,
        )
    )
    print(
        _trow(
            "Std",
            f"{np.std(fold_train_costs):.6f}",
            f"{np.std(fold_val_costs):.6f}",
            "",
            widths=w3,
        )
    )
    print(_tsep(w3))

    print(f"\n  Best Fold {best_fold + 1} — Parameters Mapped to Features:")
    w3b = [6, 30, 16]
    print(_trow("θ", "Feature / Variable", "Value", widths=w3b))
    print(_tsep(w3b))
    for i, (lbl, val) in enumerate(zip(param_labels, fold_thetas[best_fold].flatten())):
        print(_trow(f"θ{i}", lbl, f"{val:+.6f}", widths=w3b))
    print(_tsep(w3b))

    # ── TABLE 4: Part A vs Part B — parameter comparison ────────────────────
    print("\n\n" + "=" * 70)
    print("  TABLE 4 — PART A vs PART B: Parameter Comparison (CCPP)")
    print("=" * 70)
    w4 = [6, 28, 18, 18, 12]
    print(
        _trow(
            "θ",
            "Feature",
            "Part A (80/20 split)",
            f"Part B (Fold {best_fold+1})",
            "Δ Diff",
            widths=w4,
        )
    )
    print(_tsep(w4))
    for i, (lbl, va, vb) in enumerate(
        zip(param_labels, theta_a.flatten(), fold_thetas[best_fold].flatten())
    ):
        print(
            _trow(
                f"θ{i}", lbl, f"{va:+.6f}", f"{vb:+.6f}", f"{abs(va-vb):.4f}", widths=w4
            )
        )
    print(_tsep(w4))
    print(
        f"  Val Cost → Part A: {vc_n[-1]:.6f}   |   "
        f"Part B Best Fold: {fold_val_costs[best_fold]:.6f}"
    )

    # ── TABLE 5: Part C — polynomial degree comparison ───────────────────────
    print("\n\n" + "=" * 70)
    print("  TABLE 5 — PART C: Polynomial Regression Comparison (data_02b.csv)")
    print("=" * 70)
    best_d_c = min(poly_results, key=lambda d: poly_results[d]["val_error"])
    w5 = [10, 16, 16, 12, 12]
    print(
        _trow("Degree d", "Train Cost J", "Val Cost J", "# Params", "Status", widths=w5)
    )
    print(_tsep(w5))
    for d in [1, 2, 3]:
        r = poly_results[d]
        status = "★ BEST" if d == best_d_c else ""
        print(
            _trow(
                f"  d = {d}",
                f"{r['train_error']:.6f}",
                f"{r['val_error']:.6f}",
                d + 1,
                status,
                widths=w5,
            )
        )
    print(_tsep(w5))

    for d in [1, 2, 3]:
        r = poly_results[d]
        marker = " ★ BEST" if d == best_d_c else ""
        print(f"\n  d = {d}{marker} — Learned Parameters:")
        poly_labels = ["Intercept (θ₀)"] + [
            f"x^{p} coefficient (θ{p})" for p in range(1, d + 1)
        ]
        w5b = [6, 30, 16]
        print(_trow("θ", "Term", "Value", widths=w5b))
        print(_tsep(w5b))
        for i, (lbl, val) in enumerate(zip(poly_labels, r["theta"].flatten())):
            print(_trow(f"θ{i}", lbl, f"{val:+.6f}", widths=w5b))
        print(_tsep(w5b))

    # ── TABLE 6: Overall summary ──────────────────────────────────────────────
    print("\n\n" + "=" * 70)
    print("  TABLE 6 — OVERALL SUMMARY: All Parts Compared")
    print("=" * 70)
    w6 = [28, 12, 12, 22]
    print(_trow("Experiment", "Train J", "Val J", "Best Config", widths=w6))
    print(_tsep(w6))
    print(
        _trow(
            "Part A (no normalization)",
            f"{tc_nn[-1]:.4f}",
            f"{vc_nn[-1]:.4f}",
            "lr=1e-7, iters=1500",
            widths=w6,
        )
    )
    print(
        _trow(
            "Part A (with normalization)",
            f"{tc_n[-1]:.4f}",
            f"{vc_n[-1]:.4f}",
            "lr=0.01, iters=1500  ★",
            widths=w6,
        )
    )
    print(
        _trow(
            f"Part B (best fold {best_fold+1}/5, CV)",
            f"{fold_train_costs[best_fold]:.4f}",
            f"{fold_val_costs[best_fold]:.4f}",
            "lr=0.01, iters=1500",
            widths=w6,
        )
    )
    print(
        _trow(
            "Part B (CV mean across 5 folds)",
            f"{np.mean(fold_train_costs):.4f}",
            f"{np.mean(fold_val_costs):.4f}",
            f"±{np.std(fold_val_costs):.4f} std",
            widths=w6,
        )
    )
    for d in [1, 2, 3]:
        r = poly_results[d]
        marker = " ★" if d == best_d_c else ""
        print(
            _trow(
                f"Part C poly d={d}",
                f"{r['train_error']:.4f}",
                f"{r['val_error']:.4f}",
                f"lr=0.01, iters=2000{marker}",
                widths=w6,
            )
        )
    print(_tsep(w6))
    print(f"\n  ★ = best configuration in each part")


# ═══════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════


if __name__ == "__main__":
    theta_a, norm_params_a = run_part_a()
    theta_b = run_part_b()
    poly_results = run_part_c()

    print_result_tables(theta_a, theta_b, poly_results)

    print("\n" + "=" * 65)
    print("  ALL PARTS COMPLETE")
    print("=" * 65)
